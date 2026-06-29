import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
import io
import math

st.set_page_config(page_title="Plotting Jadwal Bikom", page_icon="📅", layout="wide")

st.title("📅 Aplikasi Plotting Jadwal Bimbingan Komunal (Bikom)")
st.markdown("Aplikasi ini dibuat berdasarkan PANDUAN PLOTTING JADWAL BIKOM untuk memudahkan proses penjadwalan secara otomatis.")

# === SIDEBAR UNTUK PENGATURAN ===
st.sidebar.header("⚙️ Pengaturan Parameter")
DURASI_PER_TIM = st.sidebar.number_input("Durasi Per Tim (Menit)", min_value=1, value=20, step=5)
MAKS_TIM_PER_DOSEN_PER_SESI = st.sidebar.number_input("Maksimal Tim / Dosen / Sesi", min_value=1, value=12, step=1)
MAX_TABLE_PER_ROW = st.sidebar.number_input("Jumlah Tabel Per Baris (Excel)", min_value=1, value=6, step=1)
OUTPUT_FILE_NAME = st.sidebar.text_input("Nama File Output", value="jadwal_bikom_final.xlsx")

if not OUTPUT_FILE_NAME.endswith(".xlsx"):
    OUTPUT_FILE_NAME += ".xlsx"

# === UPLOAD FILE EXCEL ===
st.header("📂 Upload Data Excel")
col1, col2 = st.columns(2)

with col1:
    st.subheader("Data Tim")
    file_tim = st.file_uploader("Upload File Excel Tim (Nama Ketua, NRP, Bidang PKM)", type=["xlsx"], key="tim")

with col2:
    st.subheader("Data Dosen")
    file_dosen = st.file_uploader("Upload File Excel Dosen (Nama Lengkap, Bidang PKM, Jam Kesediaan..., Lokasi)", type=["xlsx"], key="dosen")

if file_tim and file_dosen:
    try:
        df_tim = pd.read_excel(file_tim)
        df_dosen = pd.read_excel(file_dosen)
        
        st.success("✅ File berhasil diupload dan dibaca!")
        
        col3, col4 = st.columns(2)
        with col3:
            st.info(f"👨‍🎓 **Jumlah Tim:** {len(df_tim)}")
        with col4:
            st.info(f"👨‍🏫 **Jumlah Dosen:** {len(df_dosen)}")
            
        st.divider()
        
        # === KONFIGURASI HARI BIMBINGAN ===
        st.header("🗓️ Konfigurasi Hari Bimbingan")
        st.markdown("Pilih kolom di file Excel Dosen yang merepresentasikan **Jam Kesediaan** untuk masing-masing hari.")
        
        # Cari otomatis kolom yang mengandung kata "Jam Kesediaan"
        jam_columns = [col for col in df_dosen.columns if "Jam Kesediaan" in col or "Senin" in col or "Selasa" in col or "Rabu" in col or "Kamis" in col or "Jumat" in col or "Sabtu" in col or "Minggu" in col]
        all_columns = df_dosen.columns.tolist()
        
        hari_list = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        
        HARI_BIMBINGAN = {}
        
        num_hari = st.number_input("Jumlah Hari Bimbingan", min_value=1, max_value=7, value=len(jam_columns) if jam_columns else 2, step=1)
        
        for i in range(int(num_hari)):
            c1, c2 = st.columns(2)
            with c1:
                hari_name = st.selectbox(f"Nama Hari ke-{i+1}", hari_list, index=i if i < len(hari_list) else 0, key=f"hari_name_{i}")
            with c2:
                default_idx = 0
                for idx, opt in enumerate(all_columns):
                    if hari_name.lower() in opt.lower():
                        default_idx = idx
                        break
                
                kolom_excel = st.selectbox(f"Kolom Kesediaan Excel ({hari_name})", all_columns, index=default_idx, key=f"hari_col_{i}")
            
            if kolom_excel:
                HARI_BIMBINGAN[hari_name] = kolom_excel
            
        st.divider()
        
        # === TOMBOL PLOTTING ===
        if st.button("🚀 Mulai Plotting Jadwal", type="primary", use_container_width=True):
            with st.spinner("Sedang memproses plotting..."):
                
                # FUNGSI JAM
                def buat_slot(jam_mulai, jam_selesai, durasi):
                    slot = []
                    start = datetime.strptime(jam_mulai, "%H.%M")
                    end = datetime.strptime(jam_selesai, "%H.%M")
                    
                    while start + timedelta(minutes=durasi) <= end:
                        slot.append(
                            f"{start.strftime('%H:%M')}-{(start + timedelta(minutes=durasi)).strftime('%H:%M')}"
                        )
                        start += timedelta(minutes=durasi)
                    return slot

                def parse_jam_per_sesi(jam_text):
                    if pd.isna(jam_text):
                        return []
                    
                    jam_text = str(jam_text).strip()
                    if jam_text.lower() in ["tidak bersedia", "nan", "-", ""]:
                        return []
                    
                    sesi_list = []
                    for bagian in jam_text.split(","):
                        bagian = bagian.strip().replace("–", "-").replace("—", "-")
                        if "-" not in bagian:
                            continue
                            
                        try:
                            mulai, selesai = bagian.split("-")
                            mulai = mulai.strip()
                            selesai = selesai.strip()
                            sesi_list.append({
                                "range": f"{mulai} - {selesai}",
                                "slots": buat_slot(mulai, selesai, DURASI_PER_TIM)
                            })
                        except Exception:
                            continue
                    return sesi_list

                def jam_mulai_sesi(sesi_range):
                    mulai = sesi_range.split(" - ")[0]
                    return datetime.strptime(mulai, "%H.%M")

                def hitung_sisa_jam(sesi_range, slot_terpakai):
                    mulai, selesai = sesi_range.split(" - ")
                    selesai_dt = datetime.strptime(selesai, "%H.%M")
                    
                    if not slot_terpakai:
                        return sesi_range
                        
                    last_end = slot_terpakai[-1].split("-")[1]
                    last_end_dt = datetime.strptime(last_end, "%H:%M")
                    
                    if last_end_dt >= selesai_dt:
                        return None
                        
                    return f"{last_end_dt.strftime('%H:%M')} - {selesai}"
                
                try:
                    # PROSES DATA DOSEN
                    slot_dosen = {}
                    
                    for _, row in df_dosen.iterrows():
                        nama = str(row.get("Nama Lengkap", "Tanpa Nama")).strip()
                        if pd.isna(nama) or nama == "nan" or not nama:
                            continue
                            
                        bidang_str = str(row.get("Bidang PKM", ""))
                        bidang = [b.strip() for b in bidang_str.split(",") if b.strip()]
                        lokasi = str(row.get("Lokasi", ""))
                        if lokasi == "nan": lokasi = "-"
                        
                        slot_dosen[nama] = {
                            "bidang": bidang,
                            "lokasi": lokasi,
                            "jadwal": {}
                        }
                        
                        for hari, kolom in HARI_BIMBINGAN.items():
                            if kolom not in row:
                                continue
                                
                            sesi = parse_jam_per_sesi(row[kolom])
                            if sesi:
                                slot_dosen[nama]["jadwal"][hari] = sesi

                    # PENJADWALAN TIM
                    hasil = []
                    jumlah_tim_dosen = {}
                    
                    for _, tim in df_tim.iterrows():
                        ketua = str(tim.get("Nama Ketua", "Tanpa Nama")).strip()
                        if pd.isna(ketua) or ketua == "nan" or not ketua:
                            continue
                            
                        nrp = str(tim.get("NRP", ""))
                        if nrp == "nan": nrp = ""
                        
                        bidang_tim = str(tim.get("Bidang PKM", "")).strip()
                        assigned = False
                        
                        for dosen, data in slot_dosen.items():
                            if assigned or bidang_tim not in data["bidang"]:
                                continue
                                
                            jumlah_tim_dosen.setdefault(dosen, {})
                            
                            for hari, sesi_list in data["jadwal"].items():
                                jumlah_tim_dosen[dosen].setdefault(hari, {})
                                
                                for sesi in sesi_list:
                                    sesi_range = sesi["range"]
                                    jumlah_tim_dosen[dosen][hari].setdefault(sesi_range, 0)
                                    
                                    if jumlah_tim_dosen[dosen][hari][sesi_range] >= MAKS_TIM_PER_DOSEN_PER_SESI:
                                        continue
                                        
                                    if sesi["slots"]:
                                        jam = sesi["slots"].pop(0)
                                        
                                        hasil.append({
                                            "Dosen": dosen,
                                            "Hari": hari,
                                            "Sesi": sesi_range,
                                            "Jam": jam,
                                            "Ketua": ketua,
                                            "NRP": nrp,
                                            "Bidang": bidang_tim,
                                            "Lokasi": data["lokasi"]
                                        })
                                        
                                        jumlah_tim_dosen[dosen][hari][sesi_range] += 1
                                        assigned = True
                                        break
                                if assigned:
                                    break
                            if assigned:
                                break
                                    
                    # GROUPING
                    grouped = {}
                    for h in hasil:
                        grouped.setdefault(h["Hari"], {})
                        grouped[h["Hari"]].setdefault(h["Dosen"], {})
                        grouped[h["Hari"]][h["Dosen"]].setdefault(h["Sesi"], [])
                        grouped[h["Hari"]][h["Dosen"]][h["Sesi"]].append(h)

                    # SETUP EXCEL
                    wb = Workbook()
                    del wb["Sheet"]
                    
                    center = Alignment(horizontal="center", vertical="center")
                    border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    fill_dosen = PatternFill("solid", fgColor="3d85c6")
                    fill_hari = PatternFill("solid", fgColor="9fc5e8")
                    fill_lokasi = PatternFill("solid", fgColor="fff2cc")
                    
                    TABLE_WIDTH = 6
                    HEADER_HEIGHT = 4
                    MAX_BARIS = MAKS_TIM_PER_DOSEN_PER_SESI
                    TABLE_HEIGHT = HEADER_HEIGHT + MAX_BARIS
                    
                    # SHEET PER HARI
                    for hari in sorted(grouped.keys()):
                        ws = wb.create_sheet(hari)
                        table_idx = 0
                        
                        daftar_tabel = []
                        for dosen, sesi_data in grouped[hari].items():
                            for sesi_range, items in sesi_data.items():
                                daftar_tabel.append((sesi_range, dosen, items))
                                
                        daftar_tabel.sort(key=lambda x: jam_mulai_sesi(x[0]))
                        
                        for sesi_range, dosen, items in daftar_tabel:
                            start_col = 1 + (table_idx % MAX_TABLE_PER_ROW) * TABLE_WIDTH
                            start_row = 1 + (table_idx // MAX_TABLE_PER_ROW) * (TABLE_HEIGHT + 1)
                            
                            ws.merge_cells(start_row=start_row, start_column=start_col,
                                           end_row=start_row, end_column=start_col+4)
                            ws.cell(start_row, start_col, dosen).fill = fill_dosen
                            ws.cell(start_row, start_col).alignment = center
                            
                            ws.merge_cells(start_row=start_row+1, start_column=start_col,
                                           end_row=start_row+1, end_column=start_col+4)
                            ws.cell(start_row+1, start_col, f"{hari}, {sesi_range}").fill = fill_hari
                            ws.cell(start_row+1, start_col).alignment = center
                            
                            ws.merge_cells(start_row=start_row+2, start_column=start_col,
                                           end_row=start_row+2, end_column=start_col+4)
                            ws.cell(start_row+2, start_col, items[0]["Lokasi"]).fill = fill_lokasi
                            ws.cell(start_row+2, start_col).alignment = center
                            
                            headers = ["No", "KSN", "Jam", "Nama Ketua", "Bidang"]
                            for i, h in enumerate(headers):
                                ws.cell(start_row+3, start_col+i, h).alignment = center
                                
                            ws.merge_cells(
                                start_row=start_row+4,
                                start_column=start_col+1,
                                end_row=start_row+4+MAX_BARIS-1,
                                end_column=start_col+1
                            )
                            
                            for i in range(MAX_BARIS):
                                r = start_row + 4 + i
                                ws.cell(r, start_col, i+1)
                                if i < len(items):
                                    ws.cell(r, start_col+2, items[i]["Jam"])
                                    ws.cell(r, start_col+3, items[i]["Ketua"])
                                    ws.cell(r, start_col+4, items[i]["Bidang"])
                                    
                            for r in range(start_row, start_row + 4 + MAX_BARIS):
                                for c in range(start_col, start_col + 5):
                                    ws.cell(r, c).border = border
                                    
                            table_idx += 1

                    # SHEET TIM BELUM TERPLOT
                    tim_terplot = set(str(h["NRP"]) for h in hasil if h["NRP"])
                    ws_tim = wb.create_sheet("Tim Belum Terplot")
                    ws_tim.append(["NRP", "Nama Ketua", "Bidang PKM"])
                    
                    for _, row in df_tim.iterrows():
                        row_nrp = str(row.get("NRP", ""))
                        if row_nrp == "nan": row_nrp = ""
                        
                        if row_nrp not in tim_terplot:
                            ws_tim.append([row_nrp, str(row.get("Nama Ketua", "")), str(row.get("Bidang PKM", ""))])
                            
                    # SHEET DOSEN BELUM TERPLOT
                    ws_dosen = wb.create_sheet("Dosen Belum Terplot")
                    ws_dosen.append(["Nama Dosen", "Hari", "Sesi Jam", "Jam Tidak Terplot", "Lokasi"])
                    
                    for dosen, data in slot_dosen.items():
                        for hari, sesi_list in data["jadwal"].items():
                            for sesi in sesi_list:
                                slot_terpakai = [
                                    h["Jam"] for h in hasil
                                    if h["Dosen"] == dosen and h["Hari"] == hari and h["Sesi"] == sesi["range"]
                                ]
                                sisa = hitung_sisa_jam(sesi["range"], slot_terpakai)
                                if sisa:
                                    ws_dosen.append([dosen, hari, sesi["range"], sisa, data["lokasi"]])
                                    
                    # SHEET REKAP
                    ws_rekap = wb.create_sheet("Rekap")
                    ws_rekap.append(["REKAP JADWAL BIMBINGAN KOMUNAL"])
                    ws_rekap.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
                    ws_rekap.cell(1, 1).alignment = center
                    
                    df_hasil = pd.DataFrame(hasil)
                    
                    ws_rekap.append([])
                    ws_rekap.append(["Jumlah Tim per Hari per Bidang PKM"])
                    if not df_hasil.empty:
                        pivot = pd.pivot_table(df_hasil, index="Hari", columns="Bidang",
                                               values="Ketua", aggfunc="count", fill_value=0)
                        ws_rekap.append(["Hari"] + list(pivot.columns))
                        for hari, values in pivot.iterrows():
                            ws_rekap.append([hari] + list(values))
                            
                        ws_rekap.append([])
                        ws_rekap.append(["Jumlah Tim per Hari"])
                        ws_rekap.append(["Hari", "Jumlah Tim"])
                        for hari, count in df_hasil.groupby("Hari")["Ketua"].count().items():
                            ws_rekap.append([hari, count])
                            
                        ws_rekap.append([])
                        ws_rekap.append(["Jumlah Dosen per Hari"])
                        ws_rekap.append(["Hari", "Jumlah Dosen"])
                        for hari, count in df_hasil.groupby("Hari")["Dosen"].nunique().items():
                            ws_rekap.append([hari, count])
                            
                        ws_rekap.append([])
                        ws_rekap.append(["Jumlah Tim per Dosen"])
                        ws_rekap.append(["Nama Dosen", "Jumlah Tim"])
                        for dosen, count in df_hasil.groupby("Dosen")["Ketua"].count().items():
                            ws_rekap.append([dosen, count])
                    else:
                        ws_rekap.append(["Belum ada tim yang terplot."])

                    # Save Excel to memory
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.success(f"🎉 Plotting berhasil! {len(hasil)} tim berhasil diplot.")
                    
                    st.download_button(
                        label="⬇️ Download Hasil Plotting Excel",
                        data=excel_buffer,
                        file_name=OUTPUT_FILE_NAME,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                except Exception as e:
                    import traceback
                    st.error(f"❌ Terjadi kesalahan saat memproses data: {str(e)}")
                    with st.expander("Detail Error"):
                        st.code(traceback.format_exc())

    except Exception as e:
        st.error(f"Gagal membaca file Excel. Pastikan format file sesuai. Error: {str(e)}")
else:
    st.info("Silakan upload kedua file Excel (Data Tim & Data Dosen) di atas untuk memulai.")
