import streamlit as st
import pandas as pd
import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side

st.title("Generator Jadwal Bimbingan Komunal")

# 1. INPUT FILE EXCEL (Menggantikan files.upload() dari Colab)
file_tim = st.file_uploader("Upload File Excel Data Tim", type=['xlsx'])
file_dosen = st.file_uploader("Upload File Excel Data Dosen", type=['xlsx'])

if file_tim and file_dosen:
    df_tim = pd.read_excel(file_tim)
    df_dosen = pd.read_excel(file_dosen)
    
    st.success("Input berhasil dibaca")
    st.write(f"Jumlah Tim: {len(df_tim)}")
    st.write(f"Jumlah Dosen: {len(df_dosen)}")

    if st.button("Buat Jadwal"):
        # PARAMETER UTAMA
        DURASI_PER_TIM = 20                     #======================== INI JANGAN LUPA DISESUAIKAN (DALAM MENIT) ==============================================
        MAKS_TIM_PER_DOSEN_PER_SESI = 12        #======================== INI JANGAN LUPA DISESUAIKAN ==============================================
        OUTPUT_FILE = "jadwal_bikom2_liga.xlsx" #======================== INI JANGAN LUPA DISESUAIKAN ==============================================

        # HARI BIKOM
        HARI_BIMBINGAN = {
            "Senin": "Jam Kesediaan pada Senin, 26 Januari 2026", #======================== INI JANGAN LUPA DISESUAIKAN ==============================================
            "Selasa": "Jam Kesediaan pada Selasa, 27 Januari 2026", #======================== INI JANGAN LUPA DISESUAIKAN (BISA DITAMBAH/DIKURANGI) ==============================================
            "Rabu": "Jam Kesediaan pada Rabu, 28 Januari 2026", #======================== INI JANGAN LUPA DISESUAIKAN ==============================================
            "Kamis": "Jam Kesediaan pada Kamis, 29 Januari 2026", #======================== INI JANGAN LUPA DISESUAIKAN (BISA DITAMBAH/DIKURANGI) ==============================================
            "Jumat": "Jam Kesediaan pada Jumat, 30 Januari 2026", #======================== INI JANGAN LUPA DISESUAIKAN ==============================================
        }

        # FUNGSI JAM
        def buat_slot(jam_mulai, jam_selesai, durasi):
            slot = []
            start = datetime.strptime(jam_mulai, "%H.%M")
            end = datetime.strptime(jam_selesai, "%H.%M")

            while start + timedelta(minutes=durasi) <= end:
                slot.append(
                    f"{start.strftime('%H:%M')}-{(start + timedelta(minutes=durasi)).strftime('%H:%M')}"
                )
                start += timedelta(minutes=durasi)
            return slot

        def parse_jam_per_sesi(jam_text):
            if pd.isna(jam_text):
                return []

            jam_text = str(jam_text).strip()
            if jam_text == "Tidak Bersedia":
                return []

            sesi_list = []
            for bagian in jam_text.split(","):
                bagian = bagian.strip().replace("–", "-").replace("—", "-")
                if " - " not in bagian:
                    continue

                mulai, selesai = bagian.split(" - ")
                sesi_list.append({
                    "range": f"{mulai} - {selesai}",
                    "slots": buat_slot(mulai, selesai, DURASI_PER_TIM)
                })
            return sesi_list

        def jam_mulai_sesi(sesi_range):
            mulai = sesi_range.split(" - ")[0]
            return datetime.strptime(mulai, "%H.%M")

        def hitung_sisa_jam(sesi_range, slot_terpakai):
            mulai, selesai = sesi_range.split(" - ")
            selesai_dt = datetime.strptime(selesai, "%H.%M")

            if not slot_terpakai:
                return sesi_range

            last_end = slot_terpakai[-1].split("-")[1]
            last_end_dt = datetime.strptime(last_end, "%H:%M")

            if last_end_dt >= selesai_dt:
                return None

            return f"{last_end_dt.strftime('%H:%M')} - {selesai}"

        # PROSES DATA DOSEN
        slot_dosen = {}
        for _, row in df_dosen.iterrows():
            nama = row["Nama Lengkap"]
            bidang = [b.strip() for b in row["Bidang PKM"].split(",")]
            lokasi = row["Lokasi"]

            slot_dosen[nama] = {
                "bidang": bidang,
                "lokasi": lokasi,
                "jadwal": {}
            }

            for hari, kolom in HARI_BIMBINGAN.items():
                if kolom not in row:
                    continue

                sesi = parse_jam_per_sesi(row[kolom])
                if sesi:
                    slot_dosen[nama]["jadwal"][hari] = sesi

        # PENJADWALAN TIM
        hasil = []
        jumlah_tim_dosen = {}

        for _, tim in df_tim.iterrows():
            ketua = tim["Nama Ketua"]
            nrp = tim["NRP"]  # Pastikan kolom ini ada di Excel agar tidak memicu KeyError
            bidang_tim = tim["Bidang PKM"]
            assigned = False

            for dosen, data in slot_dosen.items():
                if assigned or bidang_tim not in data["bidang"]:
                    continue

                jumlah_tim_dosen.setdefault(dosen, {})

                for hari, sesi_list in data["jadwal"].items():
                    jumlah_tim_dosen[dosen].setdefault(hari, {})

                    for sesi in sesi_list:
                        sesi_range = sesi["range"]
                        jumlah_tim_dosen[dosen][hari].setdefault(sesi_range, 0)

                        if jumlah_tim_dosen[dosen][hari][sesi_range] >= MAKS_TIM_PER_DOSEN_PER_SESI:
                            continue

                        if sesi["slots"]:
                            jam = sesi["slots"].pop(0)

                            hasil.append({
                                "Dosen": dosen,
                                "Hari": hari,
                                "Sesi": sesi_range,
                                "Jam": jam,
                                "Ketua": ketua,
                                "NRP": nrp,
                                "Bidang": bidang_tim,
                                "Lokasi": data["lokasi"]
                            })

                            jumlah_tim_dosen[dosen][hari][sesi_range] += 1
                            assigned = True
                            break
                    if assigned:
                        break

        # GROUPING
        grouped = {}
        for h in hasil:
            grouped.setdefault(h["Hari"], {})
            grouped[h["Hari"]].setdefault(h["Dosen"], {})
            grouped[h["Hari"]][h["Dosen"]].setdefault(h["Sesi"], [])
            grouped[h["Hari"]][h["Dosen"]][h["Sesi"]].append(h)

        # SETUP EXCEL
        wb = Workbook()
        del wb["Sheet"]

        center = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        fill_dosen = PatternFill("solid", fgColor="3d85c6")
        fill_hari = PatternFill("solid", fgColor="9fc5e8")
        fill_lokasi = PatternFill("solid", fgColor="fff2cc")

        TABLE_WIDTH = 6
        HEADER_HEIGHT = 4
        MAX_BARIS = MAKS_TIM_PER_DOSEN_PER_SESI
        TABLE_HEIGHT = HEADER_HEIGHT + MAX_BARIS
        MAX_TABLE_PER_ROW = 6

        # SHEET PER HARI
        for hari in sorted(grouped.keys()):
            ws = wb.create_sheet(hari)
            table_idx = 0

            daftar_tabel = []
            for dosen, sesi_data in grouped[hari].items():
                for sesi_range, items in sesi_data.items():
                    daftar_tabel.append((sesi_range, dosen, items))

            daftar_tabel.sort(key=lambda x: jam_mulai_sesi(x[0]))

            for sesi_range, dosen, items in daftar_tabel:
                start_col = 1 + (table_idx % MAX_TABLE_PER_ROW) * TABLE_WIDTH
                start_row = 1 + (table_idx // MAX_TABLE_PER_ROW) * (TABLE_HEIGHT + 1)

                ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+4)
                ws.cell(start_row, start_col, dosen).fill = fill_dosen
                ws.cell(start_row, start_col).alignment = center

                ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+4)
                ws.cell(start_row+1, start_col, f"{hari}, {sesi_range}").fill = fill_hari
                ws.cell(start_row+1, start_col).alignment = center

                ws.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=start_col+4)
                ws.cell(start_row+2, start_col, items[0]["Lokasi"]).fill = fill_lokasi
                ws.cell(start_row+2, start_col).alignment = center

                headers = ["No", "KSN", "Jam", "Nama Ketua", "Bidang"]
                for i, h in enumerate(headers):
                    ws.cell(start_row+3, start_col+i, h).alignment = center

                ws.merge_cells(start_row=start_row+4, start_column=start_col+1, end_row=start_row+4+MAX_BARIS-1, end_column=start_col+1)

                for i in range(MAX_BARIS):
                    r = start_row + 4 + i
                    ws.cell(r, start_col, i+1)
                    if i < len(items):
                        ws.cell(r, start_col+2, items[i]["Jam"])
                        ws.cell(r, start_col+3, items[i]["Ketua"])
                        ws.cell(r, start_col+4, items[i]["Bidang"])

                for r in range(start_row, start_row + 4 + MAX_BARIS):
                    for c in range(start_col, start_col + 5):
                        ws.cell(r, c).border = border

                table_idx += 1

        # SHEET TIM BELUM TERPLOT
        tim_terplot = set(h["NRP"] for h in hasil)
        ws_tim = wb.create_sheet("Tim Belum Terplot")
        ws_tim.append(["NRP", "Nama Ketua", "Bidang PKM"])

        for _, row in df_tim.iterrows():
            if row["NRP"] not in tim_terplot:
                ws_tim.append([row["NRP"], row["Nama Ketua"], row["Bidang PKM"]])

        # SHEET DOSEN BELUM TERPLOT
        ws_dosen = wb.create_sheet("Dosen Belum Terplot")
        ws_dosen.append(["Nama Dosen", "Hari", "Sesi Jam", "Jam Tidak Terplot", "Lokasi"])

        for dosen, data in slot_dosen.items():
            for hari, sesi_list in data["jadwal"].items():
                for sesi in sesi_list:
                    slot_terpakai = [
                        h["Jam"] for h in hasil
                        if h["Dosen"] == dosen and h["Hari"] == hari and h["Sesi"] == sesi["range"]
                    ]
                    sisa = hitung_sisa_jam(sesi["range"], slot_terpakai)
                    if sisa:
                        ws_dosen.append([dosen, hari, sesi["range"], sisa, data["lokasi"]])

        # SHEET REKAP
        ws_rekap = wb.create_sheet("Rekap")
        ws_rekap.append(["REKAP JADWAL BIMBINGAN KOMUNAL"])
        ws_rekap.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws_rekap.cell(1, 1).alignment = center

        df_hasil = pd.DataFrame(hasil)

        if not df_hasil.empty:
            ws_rekap.append([])
            ws_rekap.append(["Jumlah Tim per Hari per Bidang PKM"])
            pivot = pd.pivot_table(df_hasil, index="Hari", columns="Bidang", values="Ketua", aggfunc="count", fill_value=0)
            ws_rekap.append(["Hari"] + list(pivot.columns))
            for hari, values in pivot.iterrows():
                ws_rekap.append([hari] + list(values))

            ws_rekap.append([])
            ws_rekap.append(["Jumlah Tim per Hari"])
            ws_rekap.append(["Hari", "Jumlah Tim"])
            for hari, count in df_hasil.groupby("Hari")["Ketua"].count().items():
                ws_rekap.append([hari, count])

            ws_rekap.append([])
            ws_rekap.append(["Jumlah Dosen per Hari"])
            ws_rekap.append(["Hari", "Jumlah Dosen"])
            for hari, count in df_hasil.groupby("Hari")["Dosen"].nunique().items():
                ws_rekap.append([hari, count])

            ws_rekap.append([])
            ws_rekap.append(["Jumlah Tim per Dosen"])
            ws_rekap.append(["Nama Dosen", "Jumlah Tim"])
            for dosen, count in df_hasil.groupby("Dosen")["Ketua"].count().items():
                ws_rekap.append([dosen, count])

        # 2. OUTPUT FILE EXCEL (Menggantikan files.download() dari Colab ke Streamlit Download Button)
        output = io.BytesIO()
        wb.save(output)
        
        st.success("Penjadwalan Selesai!")
        st.download_button(
            label="Download Hasil Jadwal (Excel)",
            data=output.getvalue(),
            file_name=OUTPUT_FILE,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )